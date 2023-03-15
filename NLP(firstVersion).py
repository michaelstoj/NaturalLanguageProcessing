from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

keywords_list = [
    ["fda"],
    ["sec"],
]

articles = []
newArticles = []
################ SCRUB WEBSITE ##################

while True:
    newArticles = []
    user_input = input("Enter the number of the keyword set you want to use (1, 2, or 3) or exit: ")
    if user_input == "exit" :
        break
    selected_keywords = keywords_list[int(user_input) - 1]
    if user_input == "exit" :
        break
    else:
        
        print("GLOBENEWSWRITE")
        URL = "https://www.globenewswire.com/search/?pageSize=50"
        page = requests.get(URL)
        soup = BeautifulSoup(page.text, 'html.parser')
        
        # Extract all the <a> tags
        tags = soup.find_all('a')
        
        # Print the text between the tags
        for tag in tags:
            text = tag.text
            # Check if any of the keywords are in the text
            for keyword in selected_keywords:
                if keyword.casefold() in text.casefold(): 
                    print(text)
                    if text not in articles:
                        articles.append(text)
                        newArticles.append(text)
                    print("\n")
      
  
        ######### WebSite number 2 ###################
            
        print("BUISNESS WIRE")
        
        URL = "https://www.businesswire.com/portal/site/home/news/"
        
        # Send an HTTP request to the webpage
        page = requests.get(URL)
        soup = BeautifulSoup(page.text, 'html.parser')
        
        # Extract all the <a> tags with the "bwTitleLink" class
        tags = soup.find_all('a', class_='bwTitleLink')
        
        
        # Print the text between the tags
        for tag in tags:
            text = tag.text
            # Check if any of the keywords are in the text
            for keyword in selected_keywords:
                if keyword.casefold() in text.casefold(): 
                    print(text)
                    if text not in articles:
                        articles.append(text)
                        newArticles.append(text)
                    print("\n")


#################### website 3 ###################
                    
        print("PRNEWSWIRE")
        URL = "https://www.prnewswire.com/news-releases/news-releases-list/?page=1&pagesize=100"
        
        
        # Send an HTTP request to the webpage
        page = requests.get(URL)
        soup = BeautifulSoup(page.text, 'html.parser')
        
        # Extract all the <h3> tags
        tags = soup.find_all('h3')
        
        # Print the text between the tags
        for tag in tags:
            text = tag.text
            # Check if any of the keywords are in the text
            for keyword in selected_keywords:
                if keyword.casefold() in text.casefold(): 
                    print(text)
                    if text not in articles:
                        articles.append(text)
                        newArticles.append(text)
                    print("\n")
        
            
        print(newArticles)

    


########### THIS IS HOW YOU CREATE A WORKBOOK #############
"""
# Create a new workbook
workbook = Workbook()

# Get the active sheet
worksheet = workbook.active

# Write some data to the sheet
worksheet['A1'] = 'A1'
worksheet['B1'] = 'B1'
worksheet['A2'] = 'A2'
worksheet['B2'] = 'B2'

# Save the workbook
workbook.save('data.xlsx')
"""
###################################








######## HERE IS WHERE THE WORKBOOK DATA GOES #############
# Load the workbook
def get_data():
    
######### get the input ########
# Get input from the user

    date = "2023-01-07"
    
    stock = input("Enter your stock: ")
    workbook = load_workbook('data.xlsx')

# Get the active sheet
    worksheet = workbook.active

# Find the first empty row
    row = 1
    while True:
        if worksheet.cell(row=row, column=1).value is not None:
            row += 1
        else:
            break

# Write the data to the empty row
    worksheet.cell(row=row, column=1).value = date
    worksheet.cell(row=row, column=2).value = stock.upper()
    
# Save the workbook
    workbook.save('data.xlsx')




